from flask import Flask, render_template, request, jsonify, send_file
import csv
import json
import os
import requests
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)
app.secret_key = 'hockey_pool_2026_secret'

DATA_DIR = 'data'
os.makedirs(DATA_DIR, exist_ok=True)
DRAFT_STATE_FILE = os.path.join(DATA_DIR, 'draft_state.json')
DRAFT_CSV_FILE = os.path.join(DATA_DIR, 'draft_log.csv')
PLAYERS_FILE = os.path.join(DATA_DIR, 'players.json')
LOGO_DIR = os.path.join('static', 'logos')
os.makedirs(LOGO_DIR, exist_ok=True)
DEBUG_LOG_FILE = os.path.join(DATA_DIR, 'debug_log.txt')
CSV_FIELDS = [
    'pool_name', 'n_rounds', 'total_picks', 'pick_index', 'overall_pick', 'round', 'pick_in_round',
    'pool_player_idx', 'pool_player', 'nhl_player', 'first_name', 'last_name', 'team', 'position',
    'goals', 'assists', 'points', 'ppg'
]

# 2026 Playoff teams with team colors
TEAM_INFO = {
    'BUF': {'name': 'Buffalo Sabres',        'primary': '#003087', 'secondary': '#FCB514', 'text': '#FFFFFF'},
    'CAR': {'name': 'Carolina Hurricanes',    'primary': '#CC0000', 'secondary': '#000000', 'text': '#FFFFFF'},
    'TBL': {'name': 'Tampa Bay Lightning',    'primary': '#002868', 'secondary': '#FFFFFF', 'text': '#FFFFFF'},
    'MTL': {'name': 'Montreal Canadiens',     'primary': '#AF1E2D', 'secondary': '#192168', 'text': '#FFFFFF'},
    'PIT': {'name': 'Pittsburgh Penguins',    'primary': '#000000', 'secondary': '#FCB514', 'text': '#FCB514'},
    'PHI': {'name': 'Philadelphia Flyers',    'primary': '#F74902', 'secondary': '#000000', 'text': '#FFFFFF'},
    'BOS': {'name': 'Boston Bruins',          'primary': '#FCB514', 'secondary': '#000000', 'text': '#000000'},
    'OTT': {'name': 'Ottawa Senators',        'primary': '#C52032', 'secondary': '#C69214', 'text': '#FFFFFF'},
    'COL': {'name': 'Colorado Avalanche',     'primary': '#6F263D', 'secondary': '#236192', 'text': '#FFFFFF'},
    'DAL': {'name': 'Dallas Stars',           'primary': '#006847', 'secondary': '#8F8F8C', 'text': '#FFFFFF'},
    'MIN': {'name': 'Minnesota Wild',         'primary': '#154734', 'secondary': '#A6192E', 'text': '#FFFFFF'},
    'EDM': {'name': 'Edmonton Oilers',        'primary': '#041E42', 'secondary': '#FF4C00', 'text': '#FF4C00'},
    'ANA': {'name': 'Anaheim Ducks',          'primary': '#F47A38', 'secondary': '#B9975B', 'text': '#000000'},
    'LAK': {'name': 'Los Angeles Kings',      'primary': '#111111', 'secondary': '#A2AAAD', 'text': '#FFFFFF'},
    'UTA': {'name': 'Utah Hockey Club',       'primary': '#69B3E7', 'secondary': '#010101', 'text': '#000000'},
    'VGK': {'name': 'Vegas Golden Knights',   'primary': '#B4975A', 'secondary': '#333F48', 'text': '#000000'},
    'WSH': {'name': 'Washington Capitals',    'primary': '#041E42', 'secondary': '#C8102E', 'text': '#FFFFFF'},
    'SEA': {'name': 'Seattle Kraken',         'primary': '#001628', 'secondary': '#99D9D9', 'text': '#99D9D9'},
    'NYR': {'name': 'New York Rangers',       'primary': '#0038A8', 'secondary': '#CE1126', 'text': '#FFFFFF'},
    'NJD': {'name': 'New Jersey Devils',      'primary': '#CE1126', 'secondary': '#000000', 'text': '#FFFFFF'},
    'FLA': {'name': 'Florida Panthers',       'primary': '#041E42', 'secondary': '#C8102E', 'text': '#FFFFFF'},
    'TOR': {'name': 'Toronto Maple Leafs',    'primary': '#003E7E', 'secondary': '#FFFFFF', 'text': '#FFFFFF'},
    'DET': {'name': 'Detroit Red Wings',      'primary': '#CE1126', 'secondary': '#FFFFFF', 'text': '#FFFFFF'},
    'CBJ': {'name': 'Columbus Blue Jackets',  'primary': '#002654', 'secondary': '#CE1126', 'text': '#FFFFFF'},
    'NYI': {'name': 'New York Islanders',     'primary': '#00539B', 'secondary': '#F47D30', 'text': '#FFFFFF'},
    'WPG': {'name': 'Winnipeg Jets',          'primary': '#041E42', 'secondary': '#AC162C', 'text': '#FFFFFF'},
    'NSH': {'name': 'Nashville Predators',    'primary': '#FFB81C', 'secondary': '#041E42', 'text': '#041E42'},
    'STL': {'name': 'St. Louis Blues',        'primary': '#002F87', 'secondary': '#FCB514', 'text': '#FFFFFF'},
    'CHI': {'name': 'Chicago Blackhawks',     'primary': '#CF0A2C', 'secondary': '#000000', 'text': '#FFFFFF'},
    'CGY': {'name': 'Calgary Flames',         'primary': '#C8102E', 'secondary': '#F1BE48', 'text': '#FFFFFF'},
    'VAN': {'name': 'Vancouver Canucks',      'primary': '#00843D', 'secondary': '#00205B', 'text': '#FFFFFF'},
    'SJS': {'name': 'San Jose Sharks',        'primary': '#006D75', 'secondary': '#EA7200', 'text': '#FFFFFF'},
    'ARI': {'name': 'Arizona Coyotes',        'primary': '#8C2633', 'secondary': '#E2D6B5', 'text': '#FFFFFF'},
}


def _remote_team_logo_url(abbrev):
    abbreviation = (abbrev or '').upper()
    return f"https://assets.nhle.com/logos/nhl/svg/{abbreviation}_light.svg"

def _local_team_logo_path(abbrev):
    abbreviation = (abbrev or '').upper()
    return os.path.join(LOGO_DIR, f"{abbreviation}_light.svg")

def _local_team_logo_url(abbrev):
    abbreviation = (abbrev or '').upper()
    return f"/static/logos/{abbreviation}_light.svg"

def _ensure_team_logo(abbrev):
    local_path = _local_team_logo_path(abbrev)
    if os.path.exists(local_path):
        return local_path

    os.makedirs(LOGO_DIR, exist_ok=True)
    remote_url = _remote_team_logo_url(abbrev)
    try:
        resp = requests.get(remote_url, timeout=15)
        if resp.status_code == 200 and resp.content:
            with open(local_path, 'wb') as f:
                f.write(resp.content)
            return local_path
    except Exception:
        pass
    return None

for _code, _info in TEAM_INFO.items():
    if _ensure_team_logo(_code):
        _info['logo'] = _local_team_logo_url(_code)
    else:
        _info['logo'] = _remote_team_logo_url(_code)

# Known 2026 playoff teams as fallback
FALLBACK_PLAYOFF_TEAMS = [
    'BUF', 'CAR', 'TBL', 'MTL', 'PIT', 'PHI', 'BOS', 'OTT',
    'COL', 'DAL', 'MIN', 'EDM', 'ANA', 'LAK', 'UTA', 'VGK'
]


def get_playoff_teams():
    """Fetch current playoff teams from NHL standings API."""
    try:
        url = "https://api-web.nhle.com/v1/standings/now"
        resp = requests.get(url, timeout=10)
        data = resp.json()
        playoff_teams = []
        for team in data.get('standings', []):
            indicator = team.get('clinchIndicator', '')
            if indicator and indicator not in ('e',):
                abbrev = team.get('teamAbbrev', {})
                if isinstance(abbrev, dict):
                    abbrev = abbrev.get('default', '')
                if abbrev:
                    playoff_teams.append(abbrev)
        if len(playoff_teams) >= 16:
            return playoff_teams[:16]
    except Exception as e:
        print(f"Error fetching standings: {e}")
    return FALLBACK_PLAYOFF_TEAMS


def get_team_stats(team_abbrev):
    """Fetch current season skater stats for a team."""
    try:
        url = f"https://api-web.nhle.com/v1/club-stats/{team_abbrev}/now"
        resp = requests.get(url, timeout=10)
        if resp.status_code != 200:
            return []
        data = resp.json()
        players = []
        for skater in data.get('skaters', []):
            games = skater.get('gamesPlayed', 0)
            goals = skater.get('goals', 0)
            assists = skater.get('assists', 0)
            points = skater.get('points', 0)
            ppg = round(points / games, 3) if games > 0 else 0.0
            first = skater.get('firstName', {})
            last = skater.get('lastName', {})
            if isinstance(first, dict):
                first = first.get('default', '')
            if isinstance(last, dict):
                last = last.get('default', '')
            players.append({
                'name': f"{first} {last}",
                'first_name': first,
                'last_name': last,
                'team': team_abbrev,
                'position': skater.get('positionCode', 'N/A'),
                'games': games,
                'goals': goals,
                'assists': assists,
                'points': points,
                'ppg': ppg,
                'player_id': skater.get('playerId', 0)
            })
        return players
    except Exception as e:
        print(f"Error fetching stats for {team_abbrev}: {e}")
        return []


def load_all_players(force=False):
    """Load all playoff players, from cache or API."""
    if not force and os.path.exists(PLAYERS_FILE):
        with open(PLAYERS_FILE, 'r') as f:
            return json.load(f)

    teams = get_playoff_teams()
    all_players = []
    for team in teams:
        players = get_team_stats(team)
        all_players.extend(players)

    # Sort by points desc
    all_players.sort(key=lambda x: (x['points'], x['goals']), reverse=True)

    with open(PLAYERS_FILE, 'w') as f:
        json.dump(all_players, f, indent=2)

    return all_players


def load_draft_state():
    if os.path.exists(DRAFT_STATE_FILE):
        try:
            with open(DRAFT_STATE_FILE, 'r') as f:
                return json.load(f)
        except Exception:
            pass
    if os.path.exists(DRAFT_CSV_FILE):
        state = load_draft_state_from_csv()
        if state:
            return state
    return _empty_state()


def _empty_state():
    return {
        'pool_name': '',
        'pool_players': [],
        'draft_order': [],
        'picks': {},
        'current_pick_index': 0,
        'setup_complete': False,
        'n_rounds': 10
    }


def save_draft_state(state):
    with open(DRAFT_STATE_FILE, 'w') as f:
        json.dump(state, f, indent=2)
    write_draft_csv(state)


def append_debug_log(message):
    try:
        with open(DEBUG_LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(message + '\n')
    except Exception as e:
        print(f'Unable to write debug log: {e}')


@app.before_request
def log_request():
    append_debug_log(f"REQ {request.method} {request.path}")


def generate_snake_order(n_players, n_rounds=10):
    """Return list of pool-player indices in snake draft order."""
    order = []
    for rnd in range(n_rounds):
        if rnd % 2 == 0:
            order.extend(range(n_players))
        else:
            order.extend(range(n_players - 1, -1, -1))
    return order


def load_draft_state_from_csv():
    try:
        with open(DRAFT_CSV_FILE, 'r', encoding='utf-8', newline='') as f:
            reader = csv.DictReader(f)
            rows = [row for row in reader if row and row.get('pool_player') is not None]
            if not rows:
                return None

        pool_players = {}
        picks = {}
        draft_order = []
        n_rounds = int(rows[0].get('n_rounds', 10) or 10)

        for row in rows:
            pick_index = int(row.get('pick_index', 0))
            pool_idx = int(row.get('pool_player_idx', 0))
            pool_player = row.get('pool_player', '').strip()
            if pool_player and pool_idx not in pool_players:
                pool_players[pool_idx] = pool_player
            draft_order.append(pool_idx)

            nhl_player = row.get('nhl_player', '').strip()
            if nhl_player:
                picks[str(pick_index)] = {
                    'pick_index': pick_index,
                    'overall_pick': int(row.get('overall_pick', pick_index + 1) or (pick_index + 1)),
                    'round': int(row.get('round', 1) or 1),
                    'pick_in_round': int(row.get('pick_in_round', 1) or 1),
                    'pool_player': pool_player,
                    'pool_player_idx': pool_idx,
                    'pool_name': row.get('pool_name', '').strip(),
                    'nhl_player': nhl_player,
                    'first_name': row.get('first_name', '').strip(),
                    'last_name': row.get('last_name', '').strip(),
                    'team': row.get('team', '').strip(),
                    'position': row.get('position', '').strip(),
                    'goals': int(row.get('goals', 0) or 0),
                    'assists': int(row.get('assists', 0) or 0),
                    'points': int(row.get('points', 0) or 0),
                    'ppg': float(row.get('ppg', 0.0) or 0.0)
                }

        sorted_players = [pool_players[idx] for idx in sorted(pool_players.keys())]
        total_picks = len(rows)
        if sorted_players:
            n_players = len(sorted_players)
            expected_rounds = total_picks // n_players if n_players else n_rounds
        else:
            return None

        state = {
            'pool_name': rows[0].get('pool_name', '').strip(),
            'pool_players': sorted_players,
            'draft_order': draft_order,
            'picks': picks,
            'current_pick_index': 0,
            'setup_complete': True,
            'n_rounds': expected_rounds
        }

        # Advance to the next open pick
        for idx in range(total_picks):
            if str(idx) not in picks:
                state['current_pick_index'] = idx
                break
        else:
            state['current_pick_index'] = total_picks

        return state
    except Exception:
        return None


def write_draft_csv(state):
    try:
        if not state.get('pool_players'):
            with open(DRAFT_CSV_FILE, 'w', encoding='utf-8', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
                writer.writeheader()
            return

        pool_players = state.get('pool_players', [])
        n_players = len(pool_players)
        n_rounds = int(state.get('n_rounds', 10) or 10)
        total_picks = n_rounds * n_players
        rows = []
        for pick_index in range(total_picks):
            pool_idx = state.get('draft_order', [])[pick_index] if pick_index < len(state.get('draft_order', [])) else (pick_index % n_players)
            pool_player = pool_players[pool_idx] if pool_idx < len(pool_players) else ''
            pick = state.get('picks', {}).get(str(pick_index), {})
            rows.append({
                'pool_name': state.get('pool_name', ''),
                'n_rounds': n_rounds,
                'total_picks': total_picks,
                'pick_index': pick_index,
                'overall_pick': pick_index + 1,
                'round': (pick_index // n_players) + 1,
                'pick_in_round': (pick_index % n_players) + 1,
                'pool_player_idx': pool_idx,
                'pool_player': pool_player,
                'nhl_player': pick.get('nhl_player', ''),
                'first_name': pick.get('first_name', ''),
                'last_name': pick.get('last_name', ''),
                'team': pick.get('team', ''),
                'position': pick.get('position', ''),
                'goals': pick.get('goals', ''),
                'assists': pick.get('assists', ''),
                'points': pick.get('points', ''),
                'ppg': pick.get('ppg', '')
            })

        with open(DRAFT_CSV_FILE, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
            writer.writeheader()
            writer.writerows(rows)
    except Exception as e:
        print(f"Error writing draft CSV: {e}")


# ─────────────────────────── Routes ───────────────────────────

@app.route('/')
def index():
    append_debug_log('Serving index page')
    state = load_draft_state()
    team_info_json = json.dumps(TEAM_INFO)
    return render_template('index.html',
                           state=state,
                           team_info=TEAM_INFO,
                           team_info_json=team_info_json,
                           build_version=int(time.time()))


@app.route('/api/setup_count', methods=['POST'])
def setup_count():
    data = request.json
    n = int(data.get('n_players', 0))
    if n < 2 or n > 20:
        return jsonify({'success': False, 'error': 'Between 2 and 20 players please'})
    state = load_draft_state()
    state['pool_players'] = [''] * n
    state['setup_complete'] = False
    save_draft_state(state)
    return jsonify({'success': True, 'n_players': n})


@app.route('/api/save_pool_players', methods=['POST'])
def save_pool_players():
    data = request.json
    names = [n.strip() for n in data.get('names', [])]
    pool_name = data.get('pool_name', '').strip() or 'Main Pool'
    if not all(names):
        return jsonify({'success': False, 'error': 'All names must be filled in'})
    state = load_draft_state()
    state['pool_name'] = pool_name
    state['pool_players'] = names
    state['draft_order'] = generate_snake_order(len(names), state.get('n_rounds', 10))
    state['setup_complete'] = True
    state['picks'] = {}
    state['current_pick_index'] = 0
    save_draft_state(state)
    return jsonify({'success': True, 'draft_order': state['draft_order']})


@app.route('/api/make_pick', methods=['POST'])
def make_pick():
    data = request.json
    player_name = data.get('player_name', '').strip()
    pick_index = data.get('pick_index')

    state = load_draft_state()
    players = load_all_players()

    if not state.get('setup_complete'):
        return jsonify({'success': False, 'error': 'Draft not set up yet'})

    # Find the player
    player = next((p for p in players if p['name'] == player_name), None)
    if not player:
        return jsonify({'success': False, 'error': f'Player "{player_name}" not found'})

    n_players = len(state['pool_players'])
    total_picks = state['n_rounds'] * n_players

    if pick_index is None:
        pick_index = state['current_pick_index']

    if pick_index >= total_picks:
        return jsonify({'success': False, 'error': 'Draft is complete!'})

    # Check player not already drafted elsewhere
    picks = state.get('picks', {})
    for idx_str, pk in picks.items():
        if pk and pk['nhl_player'] == player_name and int(idx_str) != pick_index:
            return jsonify({'success': False, 'error': f'{player_name} already drafted at pick {int(idx_str)+1}'})

    pool_idx = state['draft_order'][pick_index]
    round_num = pick_index // n_players + 1
    pick_in_round = pick_index % n_players + 1

    pick_data = {
        'pick_index': pick_index,
        'overall_pick': pick_index + 1,
        'round': round_num,
        'pick_in_round': pick_in_round,
        'pool_name': state.get('pool_name', ''),
        'pool_player': state['pool_players'][pool_idx],
        'pool_player_idx': pool_idx,
        'nhl_player': player_name,
        'first_name': player['first_name'],
        'last_name': player['last_name'],
        'team': player['team'],
        'position': player['position'],
        'goals': player['goals'],
        'assists': player['assists'],
        'points': player['points'],
        'ppg': player['ppg']
    }

    state['picks'][str(pick_index)] = pick_data

    # Advance pointer to next unfilled pick
    if pick_index >= state['current_pick_index']:
        next_idx = pick_index + 1
        while str(next_idx) in state['picks'] and next_idx < total_picks:
            next_idx += 1
        state['current_pick_index'] = next_idx

    save_draft_state(state)
    return jsonify({'success': True, 'pick': pick_data, 'next_pick_index': state['current_pick_index']})


@app.route('/api/set_current_pick', methods=['POST'])
def set_current_pick():
    data = request.json
    pick_index = int(data.get('pick_index', 0))
    state = load_draft_state()
    n = len(state['pool_players'])
    total = state['n_rounds'] * n
    pick_index = max(0, min(pick_index, total - 1))
    state['current_pick_index'] = pick_index
    save_draft_state(state)
    pool_idx = state['draft_order'][pick_index]
    return jsonify({
        'success': True,
        'pick_index': pick_index,
        'pool_player': state['pool_players'][pool_idx],
        'round': pick_index // n + 1,
        'existing_pick': state['picks'].get(str(pick_index))
    })


@app.route('/api/state')
def get_state():
    return jsonify(load_draft_state())


@app.route('/api/debug_log', methods=['POST'])
def debug_log():
    data = request.get_json(silent=True) or {}
    msg = data.get('message', '').strip()
    if msg:
        append_debug_log(msg)
    return jsonify({'success': True})


@app.route('/api/players')
def get_players():
    players = load_all_players()
    state = load_draft_state()
    drafted = {pk['nhl_player'] for pk in state.get('picks', {}).values() if pk}
    pick_map = {pk['nhl_player']: {'by': pk['pool_player'], 'round': pk['round']} for pk in state.get('picks', {}).values() if pk}
    for p in players:
        p['drafted'] = p['name'] in drafted
        if p['drafted']:
            p['drafted_by'] = pick_map[p['name']]['by']
            p['round'] = pick_map[p['name']]['round']
        else:
            p['drafted_by'] = ''
            p['round'] = 0
    return jsonify(players)


@app.route('/api/refresh_players', methods=['POST'])
def refresh_players():
    players = load_all_players(force=True)
    return jsonify({'success': True, 'count': len(players)})


@app.route('/api/reset_draft', methods=['POST'])
def reset_draft():
    state = load_draft_state()
    state['picks'] = {}
    state['current_pick_index'] = 0
    save_draft_state(state)
    return jsonify({'success': True})


@app.route('/api/full_reset', methods=['POST'])
def full_reset():
    save_draft_state(_empty_state())
    return jsonify({'success': True})


@app.route('/api/export_csv')
def export_csv():
    if not os.path.exists(DRAFT_CSV_FILE):
        return jsonify({'success': False, 'error': 'Draft CSV file not found.'}), 404
    return send_file(DRAFT_CSV_FILE,
                     download_name='draft_log.csv',
                     as_attachment=True,
                     mimetype='text/csv')


@app.route('/api/load_csv', methods=['POST'])
def load_csv():
    state = load_draft_state_from_csv()
    if not state:
        return jsonify({'success': False, 'error': 'Unable to load draft from CSV.'})
    save_draft_state(state)
    return jsonify({'success': True, 'state': state})


@app.route('/api/export_excel')
def export_excel():
    players = load_all_players()

    wb = Workbook()
    ws = wb.active
    ws.title = "2026 Playoff Players"

    # Header row
    headers = ['Name', 'Team', 'Position', 'Games Played', 'Goals', 'Assists', 'Points', 'Points / Game']
    header_fill = PatternFill('solid', start_color='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    center = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    # Data rows
    alt_fill = PatternFill('solid', start_color='D9E1F2')
    normal_font = Font(name='Arial', size=10)
    center_font = Font(name='Arial', size=10)

    for row_idx, p in enumerate(players, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        ws.cell(row=row_idx, column=1, value=p['name']).font = normal_font
        ws.cell(row=row_idx, column=2, value=p['team']).alignment = center
        ws.cell(row=row_idx, column=3, value=p['position']).alignment = center
        ws.cell(row=row_idx, column=4, value=p['games']).alignment = center
        ws.cell(row=row_idx, column=5, value=p['goals']).alignment = center
        ws.cell(row=row_idx, column=6, value=p['assists']).alignment = center
        ws.cell(row=row_idx, column=7, value=p['points']).alignment = center
        # PPG as formula
        ws.cell(row=row_idx, column=8, value=f'=IF(D{row_idx}>0,G{row_idx}/D{row_idx},0)')
        ws.cell(row=row_idx, column=8).number_format = '0.000'
        ws.cell(row=row_idx, column=8).alignment = center

        if fill:
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = fill

    # Column widths
    widths = [28, 8, 10, 14, 8, 10, 10, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:H{len(players) + 1}"

    # Thin border on header
    thin = Side(style='thin', color='AAAAAA')
    for col in range(1, 9):
        ws.cell(row=1, column=col).border = Border(bottom=thin)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out,
                     download_name='2026_playoff_players.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


if __name__ == '__main__':
    app.run(debug=True, port=5050)
